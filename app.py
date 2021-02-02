import pyautogui
import tkinter as tk
import time
from tkinter import filedialog, Text
from typing import List, Any
import openpyxl as xl
import os
import pickle
import datetime
import threading
from PIL import Image
from pynput.keyboard import Controller, Key
import webbrowser
import shutil
import glob
from PIL import ImageGrab
from twilio.rest import Client
pyautogui.FAILSAFE = False


global current_meeting


def notifications(notification, remove):
    text = notifications_lable['text']
    if remove:
        if notification in text:
            ntext = text.replace('{}\n'.format(notification), '')
            notifications_lable.config(text=ntext)
    if not remove:
        if notification not in text:
            text += '{}\n'.format(notification)
        notifications_lable.config(text=text)


def float_to_time(fhours):
    ihours = int(fhours)
    timez = "%02d:%02d" % (ihours, (fhours - ihours) * 60)
    return timez


def time_to_float(t):
    t = str(t)
    lst = t.split(':')
    minz = lst[1]
    min_float = int(int(minz)*(10/6)+1)
    if min_float <= 9:
        min_str = '0{}'.format(min_float)
    else:
        min_str = str(min_float)
    hours = lst[0]
    time_string = '{}.{}'.format(hours, min_str)
    time_float = float(time_string)
    return time_float


def update_meetings_of_the_day():
    meetings_of_the_day = loop_dates(calendar_path)
    text = ''
    try:
        for meeting in meetings_of_the_day:
            name = meeting[0]
            start_time = float_to_time(time_to_float(meeting[1]))
            try:
                end_time = float_to_time(time_to_float(meeting[2]))
                duration = time_to_float(meeting[2]) - time_to_float(meeting[1])
            except TypeError:
                duration = 1.8
                end_time = None
            except IndexError:
                duration = 1.8
                end_time = None
            dur_f = float_to_time(duration)
            text += 'Meeting: {}, Start: {}, Ende: {}, Dauer: {}\n\n'.format(name, start_time, end_time, dur_f)
        notifications('Kalender Fehler', True)
    except TypeError:
        notifications('Kalender Fehler', False)
    return text

def loop_dates(path):
    days = {
        0: 'MO',
        1: 'DI',
        2: 'MI',
        3: 'DO',
        4: 'FR',
        5: 'SA',
        6: 'SO'
    }
    day_int = datetime.datetime.today().weekday()
    day = days.get(day_int)
    try:
        meetings = True
        wb = xl.open(path)
        meetings_of_the_day: List[List[Any]] = []
        for table in wb:
            for cell in table['3']:
                if str(day).upper() == str(cell.value).upper():
                    row = cell.row
                    column = cell.column
                    i = 0
                    leerstellen = 0
                    while meetings:
                        name = table.cell(row + i, column + 1).value
                        start_time = table.cell(row + i, column + 2).value
                        end_time = table.cell(row + i, column + 3).value
                        user = table.cell(row + i, column + 4).value
                        password = table.cell(row + i, column + 5).value
                        link = table.cell(row + i, column + 6).value
                        date = datetime.date.today()
                        if leerstellen == 5:
                            break
                        if user is None or password is None:
                            if link is None:
                                i += 1
                                leerstellen += 1
                                continue
                        if start_time is None:
                            i += 1
                            leerstellen += 1
                            continue
                        leerstellen = 0
                        meeting_data = [name, start_time, end_time, user, password, link, date]
                        meetings_of_the_day.append(meeting_data)
                        i += 1
        notifications('Bitte über Einstellungen Kalender festlegen', True)
        notifications('Kalender gesperrt, eventuell Kalender schließen', True)
        return meetings_of_the_day
    except FileNotFoundError:
        notifications('Bitte über Einstellungen Kalender festlegen', False)
        with open('calendar_path', 'rb') as picklefile:
            global calendar_path
            calendar_path = pickle.load(picklefile)
    except PermissionError:
        notifications('Kalender gesperrt, eventuell Kalender schließen', False)


r'''
calendar_path = r'{}\Kalender.xlsx'.format(os.path.dirname(__file__))
zoom_path = r'C:\Users\lukas\AppData\Roaming\Zoom\bin\Zoom.exe'
obs_path = r"C:\Program Files\obs-studio\bin\64bit\obs64.exe"
with open('calendar_path', 'wb') as picklefile:
    pickle.dump(calendar_path, picklefile)
#with open('zoom_path', 'wb') as picklefile:
#    pickle.dump(zoom_path, picklefile)

#'''
zoom_path = r'C:\Users\Fred\AppData\Roaming\Zoom\bin\Zoom.exe'
with open('zoom_path', 'wb') as picklefile:
    pickle.dump(zoom_path, picklefile)

global recordings_path
file_path = os.path.dirname(__file__)
with open('calendar_path', 'rb') as picklefile:
    calendar_path = pickle.load(picklefile)
with open('zoom_path', 'rb') as picklefile:
    zoom_path = pickle.load(picklefile)
with open('aufnahmen_path', 'rb') as picklefile:
    recordings_path = pickle.load(picklefile)
with open('speichern_path', 'rb') as picklefile:
    save_path = pickle.load(picklefile)


def find_save_file(name):
    path = filedialog.askopenfilename()
    raw_path = r'{}'.format(path)
    with open('{}'.format(name), 'wb') as picklefile:
        pickle.dump(raw_path, picklefile)


def find_save_path(name):
    path = filedialog.askdirectory()
    raw_path = r'{}'.format(path)
    print(raw_path)
    with open('{}'.format(name), 'wb') as picklefile:
        pickle.dump(raw_path, picklefile)


def app():
    def aktu():
        meetings_lable.config(text=str(update_meetings_of_the_day()))
        root.after(1000, aktu)
    def open_calendar():
        try:
            os.startfile(calendar_path)
        except FileNotFoundError:
            pass

    def start():
        if start_button['text'] == 'Start':
            start_button.config(text='arbeitet', bg='#%02x%02x%02x' %(137, 204, 207))
            global stop
            stop = False
            threading.Thread(target=zoom, daemon=True).start()
        elif start_button['text'] == 'arbeitet':
            stop = True
            start_button.config(text='Start', bg=bbg)

    def settings():
        size = 120
        settings_window = tk.Toplevel(root)
        settings_window.title('Einstellungen')
        settings_window.geometry('{}x{}'.format(int(size * 1.618), size))
        path_kalender_button = tk.Button(settings_window, text='Kalender finden', bg=bbg, width=25, command=lambda: find_save_file('calendar_path'))
        path_kalender_button.pack()
        path_zoom_button = tk.Button(settings_window, text='Zoom finden', bg=bbg, width=25, command=lambda: find_save_file('zoom_path'))
        path_zoom_button.pack()
        path_recordings_button = tk.Button(settings_window, text='Aufnahmen finden', bg=bbg, width=25, command=lambda: find_save_path('aufnahmen_path'))
        path_recordings_button.pack()
        path_save_button = tk.Button(settings_window, text='Speicherort festlegen', bg=bbg, width=25, command=lambda: find_save_path('speichern_path'))
        path_save_button.pack()


    lable_background_color = '#%02x%02x%02x' %(199, 221, 242)
    root = tk.Tk()
    root.title('Corona Classmate')
    root.geometry('1138x700')
    root.config(background=lable_background_color)
    bbg = '#%02x%02x%02x' % (142, 186, 229)


    global notifications_lable
    notifications_lable = tk.Label(root, text='', bg=lable_background_color)
    global notifications_lable_2
    notifications_lable_2 = tk.Label(root, text='', bg=lable_background_color)
    meetings_lable = tk.Label(root, text='test', bg=lable_background_color)
    meetings_today_lable = tk.Label(root, text='Termine heute:', bg=lable_background_color)
    meetings_lable.config(text=update_meetings_of_the_day())

    calendar_button = tk.Button(root, text='Kalender', height=3, width=12, command=open_calendar, bg=bbg)
    settins_button = tk.Button(root, text='Einstellungen', bg='White', command=settings)
    start_button = tk.Button(root, text='Start', height=3, width=12, bg=bbg, command=start)

    notifications_lable.place(x=100, y=450)
    notifications_lable_2.place(x=100, y=500)
    start_button.place(x=100, y=300)
    meetings_today_lable.place(x=700, y=20)
    meetings_lable.place(x=600, y=60)
    calendar_button.place(x=100, y=150)
    settins_button.place(x=1054, y=0)
    aktu()
    root.mainloop()


def zoom():
    while True:
        print('zoom loop')
        try:
            time.sleep(1)
            meetings = loop_dates(calendar_path)
            current_time = datetime.datetime.now().strftime('%H:%M')
            for meeting in meetings:
                start_time_float = time_to_float(meeting[1])
                start_time = float_to_time(start_time_float)
                if start_time == current_time:
                    start_meeting(meeting)
            notifications('Kalender Zugriff verweigert\n Bitte über Einstellungen Kalender festlegen', True)
            notifications('Kalender Fehler', True)
        except PermissionError:
            notifications('Kalender Zugriff verweigert\n Bitte über Einstellungen Kalender festlegen', False)
        except TypeError:
            notifications('Kalender Fehler', False)
        if stop:
            break


def start_meeting(meeting):
    identisch = False
    try:
        end_time_float = time_to_float(meeting[2])
        end_time = float_to_time(end_time_float)
    except Exception:
        start_time_float = time_to_float(meeting[1])
        end_time = float_to_time(start_time_float + 1.8)
    name = meeting[0]
    user = str(meeting[3])
    password = str(meeting[4])
    link = meeting[5]
    print(user, password)
    try:
        with open('current_meeting', 'rb') as picklefile:
            current_meeting = pickle.load(picklefile)
    except FileNotFoundError or EOFError:
        current_meeting = 0
    if current_meeting == meeting:
        identisch = True
    else:
        # über den link
        notifications('Maus nicht bewegen!', False)
        notifications('Meeting startert!', False)
        if link is not None:
            print('starte über link')
            start_stop_aufnahme(False, meeting)
            try:
                webbrowser.open(link)
            except Exception:
                start_stop_aufnahme(True, meeting)
                if user is not None and password is not None:
                    print('starte Meeting über user neu 1')
                    meeting[5] = None
                    print('test des Meetings', meeting)
                    start_meeting(meeting)
                else:
                    with open('current_meeting', 'wb') as picklefile:
                        pickle.dump(meeting, picklefile)
            time.sleep(6)
            click('meeting eröffnen link.PNG')
            time.sleep(5)
            if is_on_screen('pw nach link.PNG'):
                print('need pw for link')
                for char in password:
                    time.sleep(0.1)
                    print('char', char)
                    if char == ' ':
                        continue
                    Controller().press('{}'.format(char))
                    time.sleep(0.1)
                click('an_einem_meeting_teilnehmen_link.PNG')
            time.sleep(5)
            time_out = warteschlange(end_time)
            if time_out:
                print('time out')
                start_stop_aufnahme(True, meeting)
            if is_meeting_open(meeting, time_out) and not time_out:
                print('meeting ist offen')
                threading.Thread(target=whatsapp, args=('meeting {}  {} gestartet über link'.format(name, meeting[1]),)).start()
                threading.Thread(target=meeting_schließen, args=(end_time, meeting)).start()
            elif not time_out:
                meeting[5] = None
                start_stop_aufnahme(True, meeting)
                if user is not None and password is not None:
                    print('meeting wird über user neugestartet 2')
                    start_meeting(meeting)
            notifications('Meeting wird über user neu gestartet', True)
            notifications('Maus nicht bewegen!', True)
            notifications('Meeting startert!', True)
        else:
            print('starte über user')
            start_stop_aufnahme(False, meeting)
            print('zoom_path', zoom_path)
            os.startfile(zoom_path)
            time.sleep(5)
            click('Beitreten_über_user.PNG')
            time.sleep(3)
            print('user', user)
            for char in user:
                print('here')
                print(char)
                time.sleep(0.1)
                Controller().press('{}'.format(char))
            time.sleep(0.25)
            click('beitreten_nach_user.PNG')
            time.sleep(4)
            if is_on_screen('pw_nach_user_eingabe.PNG'):
                for char in password:
                    time.sleep(0.1)
                    if char == ' ':
                        continue
                    Controller().press('{}'.format(char))
                    print('writing pw: {}'. format(char))
                time.sleep(0.25)
                click('an_meeting_teilnehmen_nicht_angemeldet.PNG')
            time_out = warteschlange(end_time)
            print('time out,', time_out)
            if not time_out and is_meeting_open(meeting, time_out):
                print('meeting läuft')
                threading.Thread(target=whatsapp,
                                 args=('meeting {}  {} gestartet über user'.format(name, meeting[1]),)).start()
                threading.Thread(target=meeting_schließen, args=(end_time, meeting)).start()
            else:
                start_stop_aufnahme(True, meeting)
                threading.Thread(target=whatsapp,
                                 args=('meeting {}  {} konnte nicht gestartet werden'.format(name, meeting[1]),)).start()

        notifications('Maus nicht bewegen!', True)
        notifications('Meeting startert!', True)
        notifications('Meeting wird über user neu gestartet', True)
    if not identisch:
        with open('current_meeting', 'wb') as picklefile:
            pickle.dump(meeting, picklefile)


def meeting_schließen(endtime, meeting):
    print('meeting {} {} geplant schließen wird im Hintergrund ausgeführt'.format(meeting[0], meeting[6]))
    time.sleep(4)
    click('minimieren video.png')
    while True:
        time.sleep(2)
        current_time = datetime.datetime.now().strftime('%H:%M')
        if is_on_screen('vergrößern.PNG'):
            pyautogui.doubleClick(x=1150, y=540)
        if is_on_screen('Umfragen.PNG'):
            coordinates = pyautogui.locateAllOnScreen('Umfragen.PNG')
            print(coordinates)
            click('x Umfragen.PNG')
            click('Umfragen schließen.PNG')
            pyautogui.moveTo(x=1150, y=540)
        click('stummschalten.PNG')
        if is_on_screen('Hintergrund1.PNG') or is_on_screen('Hintergrund2.PNG') or is_on_screen('Hintergrund3.PNG') or is_on_screen('Hintergrund4.PNG'):
            pyautogui.doubleClick(x=1150, y=540)

        if is_on_screen('meeting vom moderator beendet.PNG'):
            start_stop_aufnahme(True, meeting)
            break
        if current_time == endtime:
            print('meeting_schließen Time_out')
            click('Als host beenden.PNG')
            notifications('Maus nicht bewegen!', False)
            notifications('Meeting wird beendet!', False)
            click('verlassen.PNG')
            time.sleep(0.5)
            click('meeting verlassen.png')
            click('für alle beenden.png')
            click('Zuweisen und verlassen.PNG')
            if not is_on_screen('obs_taskleiste.PNG'):
                pyautogui.doubleClick(x=1150, y=540)
            time.sleep(1)
            start_stop_aufnahme(True, meeting)
            notifications('Maus nicht bewegen!', True)
            notifications('Meeting wird beendet!', True)
            break


def start_stop_aufnahme(stopp_aufnahme, meeting):
    if not stopp_aufnahme:
        if not is_on_screen('obs_Aufnahme_starten.PNG'):
            click('obs_taskleiste.PNG')
            time.sleep(6)
        click('obs_Aufnahme_starten.PNG')
    if stopp_aufnahme:
        if not is_on_screen('obs_Aufnahme_beenden.PNG'):
            click('obs_taskleiste.PNG')
            time.sleep(1.5)
        click('obs_Aufnahme_beenden.PNG')
        print('aufnahme gestoppt')
        threading.Thread(target=aufnahmen_sortieren, args=(meeting, )).start()
    click('obs_taskleiste.PNG')


def click(button):
    img = Image.open(button)
    coordinates = pyautogui.locateCenterOnScreen(img, grayscale=True, confidence=.8)
    pyautogui.moveTo(coordinates)
    if coordinates is None:
        print('did not find {}'.format(button))
    if coordinates is not None:
        print('click {}'.format(button))
        pyautogui.click(coordinates)


def is_on_screen(img_path):
    img = Image.open(img_path)
    is_there = pyautogui.locateCenterOnScreen(img, grayscale=True, confidence=.8)
    if is_there is None:
        print('is not there: {}'.format(img_path))
        return False
    else:
        print('there is: {}'.format(img_path))
        return True


def warteschlange(end_time):
    time_out_meeting = False
    time.sleep(5)
    for _ in range(2):
        time.sleep(1.5)
        pyautogui.doubleClick(x=1150, y=540)
    for i in range(4):
        time.sleep(1)
        print('start warteschlange')
        print(end_time, 'endtime')
        if is_on_screen('ohne Video beitreten.PNG') or is_on_screen('warten bis hoste das meeting beginnt.PNG') or is_on_screen('warteraum_testen_Sie_das_Computeraudio.PNG') or is_on_screen('Verbindung wird aufgebaut.PNG'):
            print('warteschlange')
            click('ohne Video beitreten.PNG')
            current_time = datetime.datetime.now().strftime('%H:%M')
            while time_to_float(end_time) - 0.3 >= (time_to_float(current_time)):
                current_time = datetime.datetime.now().strftime('%H:%M')
                click('ohne Video beitreten.PNG')
                time.sleep(0.25)
                print('warte')
                if is_on_screen('verlassen.PNG') or is_on_screen('zoom läuft.PNG'):
                    return time_out_meeting
                    break
                    pass
                pass
            print('warteschlange Time out')
            time_out_meeting = True
            if is_on_screen('Bittewarten.PNG'):
                print('found bitte warten')
                img = Image.open('zoom x.PNG')
                coordinates = pyautogui.locateCenterOnScreen(img, grayscale=True, confidence=.99)
                pyautogui.moveTo(coordinates)
                if coordinates is None:
                    print('did not find{}'.format('zoom x'))
                if coordinates is not None:
                    pyautogui.click(coordinates)
                click('verlassen.PNG')
            if is_on_screen('warteraum_testen_Sie_das_Computeraudio.PNG'):
                if not is_on_screen('zoom x.PNG'):
                    time.sleep(1)
                    pyautogui.click(x=1150, y=540)
                    time.sleep(0.2)
                    pyautogui.click(x=1150, y=540)
                img = Image.open('zoom x.PNG')
                coordinates = pyautogui.locateCenterOnScreen(img, grayscale=True, confidence=.99)
                pyautogui.moveTo(coordinates)
                if coordinates is None:
                    print('did not find{}'.format('zoom x'))
                if coordinates is not None:
                    pyautogui.click(coordinates)
                click('verlassen.PNG')
    return time_out_meeting


def aufnahmen_sortieren(meeting):
    time.sleep(90)
    print('sortiere die aufnahmen')
    name = meeting[0]
    if name is None:
        name = 'ohne_Namen'
    list_of_files = glob.glob('{}\*'.format(recordings_path))
    latest_file = max(list_of_files, key=os.path.getmtime)
    datum = datetime.datetime.now().strftime(r'%Y_%m_%d__%H_%M')
    dir = r'{}\{}'.format(save_path, name)
    try:
        print('try to make dir')
        os.mkdir(dir)
        shutil.copy2(latest_file, r'{}\{}__{}.mkv'.format(dir, name, datum))
        print('Verzeichnis erstellt, Aufnahme sortiert')
    except FileExistsError:
        shutil.copy2(latest_file, r'{}\{}__{}.mkv'.format(dir, name, datum))
        print('Aufnahme erstellt')
    except PermissionError:
        time.sleep(60)
        shutil.copy2(latest_file, r'{}\{}__{}.mkv'.format(dir, name, datum))
        print('Aufnahme erstellt')


def is_meeting_open(meeting, time_out):
    meeting_läuft = True
    time.sleep(5)
    print('is meeting open?')
    if not is_on_screen('verlassen.PNG') and not is_on_screen('zoom läuft.PNG') and not is_on_screen('Als host beenden.PNG'):
        print('meeting läuft nicht')
        click('abbrechen.PNG')
        time.sleep(0.25)
        img = Image.open('zoom x.PNG')
        coordinates = pyautogui.locateCenterOnScreen(img, grayscale=True, confidence=.99)
        pyautogui.moveTo(coordinates)
        if coordinates is None:
            print('did not find{}'.format('zoom x'))
        if coordinates is not None:
            pyautogui.click(coordinates)
        time.sleep(0.25)
        click('meeting verlassen.png')
        if meeting[5] is not None and meeting[3] is not None:
            meeting[5] = None
            notifications('Meeting wird über user neu gestartet', True)
            meeting_läuft = False
            print('time out =', time_out)
            if not time_out:
                return meeting_läuft
        else:
            print('meeting konnte nicht gestartet werden')
            start_stop_aufnahme(True, meeting)
            threading.Thread(target=whatsapp,
                        args=('meeting {}  {} konnte nicht gestartet werden'.format(meeting[0], meeting[1]),)).start()
            meeting_läuft = False
            return meeting_läuft
    else:
        print('meeting_läuft Doppelclick')
        for _ in range(2):
            time.sleep(1.5)
            pyautogui.doubleClick(x=1150, y=540)
        return meeting_läuft


def whatsapp(message_body):
    account_sid = 'AC86d8ed885d29c72810499c81a60d8de5'
    auth_token = '6236d7fc9e75f22122899de6043c2e9e'
    client = Client(account_sid, auth_token)

    message = client.messages.create(
        from_='whatsapp:+14155238886',
        body=message_body,
        to='whatsapp:+4915737218444'
    )

    print(message.sid)


test_meeting = ['Metallvermittelte Synthese', datetime.time(15, 30), None, '864 9114 6721', 267552, 'https://us02web.zoom.us/j/86491146721?pwd=cnBTRk45WGhHWFV4WENPdXhrV0tXdz09', datetime.date(2020, 10, 27)]



if __name__ == '__main__':
    threading.Thread(target=app).start()

